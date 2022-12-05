import requests
import os
import logging
import json
from urllib.parse import urlparse

from dotenv import load_dotenv
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from requests.exceptions import SSLError, ConnectTimeout, ReadTimeout

CWD = os.getcwd()
BAD_REQUEST = 400

load_dotenv(os.path.abspath(os.path.join(CWD, os.pardir, ".env")))

logging.basicConfig(level=logging.INFO)

service_url = "http://127.0.0.1:5000/api/add_ev"
auth = requests.auth.HTTPBasicAuth(username=os.getenv("API_USER"), password=os.getenv("API_PASSWORD"))
parse_headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 OPR/93.0.0.0",
    "referer": "https://google.com/",
    "sec-ch-ua": '"Opera";v="93", "Not/A)Brand";v="8", "Chromium";v="107"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-au-platform": "Windows",
}
headers = {
    "content-type": "application/json"
}


def save_data_to_service_request(data: dict) -> None:
    response = requests.post(service_url, headers=headers, data=json.dumps(data, default=str), auth=auth)
    logging.info(f"Service Response -> {response.status_code}")


def parse_page(url: str) -> dict:
    response = None
    data = {
        "original_url": url,
        "status_code": BAD_REQUEST,
        "title": "",
        "domain_name": urlparse(url).netloc.replace("www.", ""),
    }
    logging.info(f"parsing url {url}")
    try:
        response = requests.get(url, headers=parse_headers, allow_redirects=True, timeout=30)
    except (ConnectTimeout, ReadTimeout):
        data["status_code"] = 408  # timeout
    except SSLError as ssl_err:
        logging.error(f"SSL Error occurred {ssl_err}. Trying to go without verifying ")

        # verify = False, for unsecure connection and getting data
        try:
            response = requests.get(url, headers=parse_headers, allow_redirects=True, timeout=30, verify=False)
        except Exception as err:
            logging.error(f"A new one error occurred {err}. Can't avoid error, would be 400...")
    except Exception as err:
        logging.error(f"Error occured {err}")

    if response:
        doc = BeautifulSoup(response.text or "", "html.parser")
        data.update({
            "status_code": response.status_code,
            "title": doc.title and doc.title.text or "",
        })
        if response.history:
            data.update({
                "final_url": response.url,
                "status_code":  response.history[-1].status_code, # status code of original page
            })
        response.close()

    logging.info(f"Parsing code -> {data['status_code']} for url - {url}")
    return data


def import_data(file_to_read: str) -> None:
    wb = load_workbook(file_to_read)
    sheet = wb.active
    # iterate through rows of xlsx file, min_row = 2 for skipping headers of table
    for row in sheet.iter_rows(min_row=2):
        visit_datetime_cell, url_cell = row  # simplify naming, as we now what is in our columns
        data_from_page = parse_page(url_cell.value)

        data_from_page["visit_date"] = visit_datetime_cell.value

        save_data_to_service_request(data_from_page)


if __name__ == "__main__":
    file_name = "Junior Strong Python Developer Test.xlsx"
    file_location = CWD  # change it due to base directory where file located
    path_to_file = os.path.join(file_location, file_name)
    import_data(path_to_file)
